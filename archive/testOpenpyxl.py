#Pour le traitement des macros du fichier excel, nous avons utilisé ce programme (premier programme).
from openpyxl import load_workbook

# Charger le fichier Excel
try:
    workbook = load_workbook(filename="alyfData.xlsm", data_only=False)
except FileNotFoundError:
    print("Le fichier Excel est introuvable.")
    exit(1)

# Accéder à la feuille nommée "DEV WEB"
if "DEV WEB" in workbook.sheetnames:
    worksheet = workbook["DEV WEB"]
else:
    print('La feuille "DEV WEB" est introuvable.')
    exit(1)

# Lire la valeur de la cellule H1
cell_value = worksheet["H1"].value
print("La valeur de la cellule H1 est:", cell_value)

# Modifier la valeur de la cellule H1
worksheet["H1"].value = "HUYNH"

# Sauvegarder les modifications dans un nouveau fichier
workbook.save("alyfDev.xlsm")

